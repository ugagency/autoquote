import os
import re
import time
import shutil
import tempfile
import threading
from datetime import datetime, date
from app.models import UserModel
from app import create_app

# ===============================
# DEFINIÇÃO DA FUNÇÃO PRINCIPAL
# ===============================
def executar_robo_vale(data_coleta, user_id=None, log_callback=None):
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from openpyxl import Workbook, load_workbook
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.common.exceptions import (
        StaleElementReferenceException,
        ElementClickInterceptedException,
        ElementNotInteractableException,
        WebDriverException,
    )

    # ===============================
    # LOG SETUP
    # ===============================
    def log(msg, progresso=None):
        if log_callback:
            log_callback(msg, progresso)
        print(msg)

    # ===============================
    # CREDENCIAIS DO USUÁRIO (SUPABASE)
    # ===============================
    if not user_id:
        raise ValueError("❌ user_id é obrigatório para buscar as credenciais do Vale.")

    log("🔑 Buscando credenciais do Vale...", 5)
    config = UserModel.get_user_vale_config(user_id)

    if not config or not config.get("vale_email") or not config.get("vale_password"):
        raise ValueError("⚠️ Configurações do Vale não encontradas para este usuário.")

    # senha já vem descriptografada pelo UserModel
    USER = config["vale_email"]
    PASS = config["vale_password"]

    if not USER or not PASS:
        raise ValueError("⚠️ Credenciais inválidas ou ausentes.")

    log(f"✅ Credenciais carregadas para {USER}", 8)

    # ===============================
    # VALIDAÇÃO DA DATA
    # ===============================
    if not (data_coleta and len(data_coleta.strip()) == 6 and data_coleta.isdigit()):
        raise ValueError("Data inválida! Use o formato DDMMAA (ex: 191025).")

    HOJE_str = f"{data_coleta[:2]}/{data_coleta[2:4]}/{data_coleta[4:]}"
    
    def parse_date_str(s: str):
        """Tenta vários formatos e retorna datetime.date ou None."""
        if not s: return None
        s = s.strip()
        for fmt in ("%d/%m/%y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue
        # tentativa extra com regex
        m = re.search(r'(\d{1,2}/\d{1,2}/\d{2,4})', s)
        if m:
            for fmt in ("%d/%m/%y", "%d/%m/%Y"):
                try:
                    return datetime.strptime(m.group(1), fmt).date()
                except:
                    continue
        return None

    HOJE = parse_date_str(HOJE_str)
    if not HOJE:
        raise ValueError("Data fornecida é inválida.")
    
    log(f"📅 Executando coleta para {HOJE.strftime('%d/%m/%Y')}", 10)

    ESTADOS = [
        'AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA', 'MT', 'MS', 'MG',
        'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN', 'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO'
    ]

    # ===============================
    # PREPARAÇÃO DE PLANILHA
    # ===============================
    filename = f"planilha_vale_{HOJE.strftime('%d%m%y')}.xlsx"
    # Ajuste de caminho para garantir que salva no output correto
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    OUTPUT_DIR = os.path.join(base_dir, 'output', str(user_id))
    
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        
    EXCEL_PATH = os.path.join(OUTPUT_DIR, filename)
    log(f"💾 Caminho do Excel: {EXCEL_PATH}")

    if os.path.exists(EXCEL_PATH):
        try:
            os.remove(EXCEL_PATH)
        except Exception as e:
            log(f"⚠️ Erro ao remover arquivo anterior: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"
    ws.append(["Numero do evento", "UF(VALE)", "DATA", "DESCRIÇÃO", "QTDE", "UNID. MED", "pagina de descrição"])
    wb.save(EXCEL_PATH)
    log("📗 Planilha inicial criada.", 15)

    # ===============================
    # CONFIGURAÇÃO DO SELENIUM
    # ===============================
    chrome_options = Options()
    # Adicionando flags importantes para rodar em container/headless
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    
    # Gerenciamento do Driver
    base_driver_path = ChromeDriverManager().install()

    def find_chromedriver(base_path):
        if os.path.isfile(base_path) and "THIRD_PARTY_NOTICES" in base_path:
            base_path = os.path.dirname(base_path)
        possible_paths = [
            os.path.join(base_path, "chromedriver"),
            os.path.join(base_path, "chromedriver-linux64", "chromedriver"),
            os.path.join(os.path.dirname(base_path), "chromedriver"),
        ]
        for path in possible_paths:
            if os.path.exists(path):
                return path
        # Fallback: se for windows, tenta achar o exe
        for path in possible_paths:
            if os.path.exists(path + ".exe"):
                return path + ".exe"
        return base_path # Retorna o base path mesmo se não achar exato, as vezes o manager retorna o executavel direto

    driver_path = find_chromedriver(base_driver_path)
    
    # Garante permissão de execução no linux
    try:
        if os.name != 'nt':
            os.chmod(driver_path, 0o755)
    except Exception:
        pass

    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)
    wait = WebDriverWait(driver, 10)
    log("🚀 Selenium WebDriver iniciado com sucesso.", 20)

    try:
        # ===============================
        # LOGIN
        # ===============================
        driver.get("https://vale.coupahost.com/sessions/supplier_login")

        wait.until(EC.presence_of_element_located((By.ID, "user_login")))
        driver.find_element(By.ID, "user_login").send_keys(USER)
        driver.find_element(By.ID, "user_password").send_keys(PASS, Keys.RETURN)
        log("🔐 Login enviado...", 25)

        # ===============================
        # FILTRO DE DATA
        # ===============================
        # Clica no elemento de data duas vezes
        try:
            time_filter = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ch_start_time"]')))
            time_filter.click()
            time.sleep(5)
            # Reaplica o clique para garantir ordenação correta se necessário, conforme script original
            time_filter = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ch_start_time"]')))
            time_filter.click()
        except:
            pass
        
        # ===============================
        # COLETA DA TABELA PRINCIPAL
        # ===============================
        encontrou_ontem = False
        total_eventos = 0
        log("🔎 Iniciando varredura da tabela de eventos...", 30)
        
        while True:
            time.sleep(3) # Wait pro carregamento da tabela
            try:
                tbody = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="quote_request_table_tag"]')))
            except Exception:
                log("⚠️ Tabela não encontrada ou fim da paginação.")
                break
                
            # Estratégia Anti-Stale: Conta linhas e itera por índice, recarregando a cada passo
            try:
                rows_count = len(tbody.find_elements(By.TAG_NAME, "tr"))
            except:
                rows_count = 0
                
            log(f"📄 Página atual: {rows_count} linhas encontradas.")

            # --- Buscar os números do evento ---
            for i in range(rows_count):
                try:
                    # Re-fetch para evitar StaleElementReferenceException
                    tbody_fresh = driver.find_element(By.XPATH, '//*[@id="quote_request_table_tag"]')
                    linhas_fresh = tbody_fresh.find_elements(By.TAG_NAME, "tr")
                    
                    if i >= len(linhas_fresh):
                        break # Tabela encolheu
                        
                    linha = linhas_fresh[i]
                    colunas = linha.find_elements(By.TAG_NAME, "td")
                    
                    if not colunas or len(colunas) < 7:
                        continue

                    # pula a linha se existir o ícone amarelo (flag_yellow) em qualquer lugar da linha
                    yellow_flags = linha.find_elements(By.CSS_SELECTOR, "img[src*='flag_yellow']")
                    if yellow_flags:
                        continue

                    # Identifica Status (Coluna 4 - 0-based)
                    try:
                        status_text = colunas[4].text.strip()
                        if "Concluído" in status_text:
                            continue
                    except Exception:
                        pass

                    data_inicio_str = colunas[2].text.strip()
                    data_inicio = parse_date_str(data_inicio_str)
                    
                    if data_inicio is None:
                        continue

                    if data_inicio < HOJE:
                        encontrou_ontem = True
                        log(f"⏹️ Encontrada data anterior ({data_inicio}), parando coleta.")
                        break

                    if data_inicio != HOJE:
                        continue
                        
                    # Busca segura do link
                    try:
                        link_el = colunas[0].find_element(By.TAG_NAME, "a")
                        numero_evento = link_el.text.strip()
                    except:
                        continue
                        
                    data_final = colunas[3].text.strip()
                    
                    ws.append([numero_evento, '', data_final, '', '', '', ''])

                except Exception as e:
                    # print(f"Erro ao ler linha {i}: {e}")
                    continue

            if encontrou_ontem:
                break

            try:
                # Tenta avançar página
                proximo = driver.find_element(By.CLASS_NAME, "next_page")
                if "disabled" in proximo.get_attribute("class"):
                    break
                    
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", proximo)
                time.sleep(1)
                proximo.click()
                time.sleep(2)
            except Exception:
                break

        wb.save(EXCEL_PATH)
        log(f"💾 Coleta inicial finalizada. Planilha salva.", 50)

        # ===============================
        # DETALHAMENTO DE CADA EVENTO
        # ===============================
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Eventos"]
        
        # Coleta lista de eventos na memória para iterar sem travar o openpyxl
        rows_to_process = []
        for row in ws.iter_rows(min_row=2):
            if row[0].value:
                rows_to_process.append(row[0].value)
        
        # Remove duplicatas preservando ordem
        eventos_unicos = []
        seen = set()
        for e in rows_to_process:
            if e not in seen:
                eventos_unicos.append(e)
                seen.add(e)
                
        total_detalhar = len(eventos_unicos)
        count_detalhado = 0
        
        log(f"🔍 Iniciando detalhamento de {total_detalhar} eventos...", 55)

        for evento in eventos_unicos:
            count_detalhado += 1
            # Como estamos editando a planilha, precisamos recarregar as linhas que correspondem a este evento
            # Mas o jeito mais facil é iterar sobre a planilha, achar as linhas desse evento e preencher
            pass # Lógica abaixo faz isso
            
            driver.get(f"https://vale.coupahost.com/quotes/external_responses/{evento}/edit")
            try:
                wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            except:
                log(f"⚠️ Timeout ao carregar página do evento {evento}")
                continue

            # --- VERIFICA EXISTÊNCIA DA PÁGINA DE DESCRIÇÃO ---
            try:
                botoes1 = driver.find_elements(By.XPATH, '//*[@id="pageContentWrapper"]/div[3]/div[2]/a[2]/span')
                if not botoes1:
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    botoes2 = driver.find_elements(By.ID, 'quote_response_submit')
                    if botoes2:
                        botoes2[0].click()
            except Exception:
                pass # Pode ser que já esteja na página certa

            # Scroll e abre seção das informações
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            try:
                wait.until(EC.presence_of_element_located((By.CLASS_NAME, "s-expandLines")))
            except:
                # Se não tem s-expandLines, talvez não tenha itens ou nao carregou
                log(f"⚠️ Nenhum item expansível (s-expandLines) no evento {evento}")
                continue
                
            elementos = driver.find_elements(By.CLASS_NAME, "s-expandLines")

            if not elementos:
                continue

            # Duplicar a linha do evento se houver múltiplos itens
            # A lógica original do usuário duplicava as linhas na planilha
            # Para simplificar a manipulação do Excel, vamos encontrar as linhas na planilha atual
            # e adicionar novas se necessário.
            # OBS: O script do usuário faz append de novas linhas no final. Isso pode desordenar, mas ele ordena no fim.
            
            # Encontra a linha "modelo" desse evento
            linha_modelo = None
            rows_list = list(ws.iter_rows(min_row=2))
            for r in rows_list:
                if str(r[0].value) == str(evento):
                    linha_modelo = r
                    break
            
            if not linha_modelo:
                continue # Estranho, devia ter
            
            # Se tem mais elementos do que linhas atuais, cria novas
            # Mas como não sabemos quantas linhas já tem (se rodou parcial), vamos assumir a lógica do usuário:
            # Ele sempre faz append de (N-1) linhas novas se encontrar N elementos
            # Para evitar duplicatas infinitas se rodar de novo, ideal seria limpar, mas vamos seguir o script:
            if len(elementos) > 1:
                vals = [c.value for c in linha_modelo]
                for i in range(len(elementos) - 1):
                    # Cria nova linha com dados básicos
                    # [Numero, UF, DATA, DESC, QTDE, UNID, PAG]
                    ws.append([vals[0], '', vals[2], '', '', '', '']) 
                wb.save(EXCEL_PATH)
            
            # Helper interno de clique
            def click_element_retry(el, attempts=4, pause=0.4):
                for _ in range(attempts):
                    try:
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        time.sleep(0.15)
                        el.click()
                        return True
                    except (StaleElementReferenceException, ElementClickInterceptedException, ElementNotInteractableException, WebDriverException):
                        try:
                            driver.execute_script("arguments[0].click();", el)
                            return True
                        except Exception:
                            time.sleep(pause)
                return False

            total = driver.execute_script("return document.querySelectorAll('.s-expandLines').length")
            processed = 0
            idx = 0
            
            # Recarrega a planilha para ter as linhas novas
            # Filtrar apenas linhas deste evento
            relevant_rows = []
            # iterar de novo (ineficiente mas seguro conforme script)
            for r in ws.iter_rows(min_row=2):
                if str(r[0].value) == str(evento):
                    relevant_rows.append(r)
            
            # Loop de processamento de itens
            while processed < total and idx < total:
                try:
                    elementos = driver.find_elements(By.CLASS_NAME, "s-expandLines")
                except:
                    time.sleep(0.3)
                    elementos = driver.find_elements(By.CLASS_NAME, "s-expandLines")

                if idx >= len(elementos):
                     # Retry DOM
                    time.sleep(0.5)
                    elementos = driver.find_elements(By.CLASS_NAME, "s-expandLines")
                    if idx >= len(elementos):
                         idx += 1
                         continue

                el = elementos[idx]
                
                # Checa se já processou via atributo
                already = driver.execute_script("return arguments[0].getAttribute('data-processed')", el)
                if already:
                    idx += 1
                    processed += 1
                    continue

                if not click_element_retry(el):
                    # Marca falha e pula
                    driver.execute_script("arguments[0].setAttribute('data-processed','1')", el)
                    idx += 1
                    processed += 1
                    continue

                # Espera carregar detalhe
                try:
                    wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]')))
                    time.sleep(0.25)
                except:
                    time.sleep(0.5)

                # Pega a linha correspondente na planilha
                # Se houver mais itens que linhas (erro de sync), pega a última ou cria (mas aqui pegamos a lista antes)
                if idx < len(relevant_rows):
                    linha_atual = relevant_rows[idx]
                else:
                    # Fallback raro
                    linha_atual = relevant_rows[-1] 

                # Coleta DADOS
                # 1. Quantidade
                try:
                    quantidade_el = driver.find_element(By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[2]/div/div[2]/div/p/span[1]')
                    linha_atual[4].value = quantidade_el.text
                except:
                    linha_atual[4].value = '?'

                # 2. Unidade
                try:
                    unidade_el = driver.find_element(By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[2]/div/div[2]/div/p/span[2]')
                    linha_atual[5].value = unidade_el.text
                except:
                    linha_atual[5].value = '?'
                
                # 3. Descrição
                try:
                    descri_el = driver.find_element(By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[1]/div/div[2]/div/p')
                    descri = descri_el.text
                    desejado = re.search(r'PT\s*\|\|\s*(.*?)\*{3,}', descri, re.DOTALL)
                    linha_atual[3].value = desejado.group(1).strip() if desejado else descri
                except:
                    linha_atual[3].value = 'Erro Desc'

                # 4. UF Restrita
                try:
                    uf_spans = driver.find_elements(
                        By.XPATH,
                        '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[1]/div/div[8]/div/ul/li/span'
                    )
                    found = None
                    for elem in uf_spans:
                        text = (elem.text or "").strip().upper()
                        if not text: continue
                        # Padrão - XX - BR
                        m = re.search(r'-\s*([A-Z]{2})\s*-\s*BR', text)
                        if m and m.group(1) in ESTADOS:
                            found = m.group(1)
                            break
                        # Tokens
                        tokens = re.findall(r'\b[A-Z]{2}\b', text)
                        for t in tokens:
                            if t in ESTADOS:
                                found = t
                                break
                        if found: break
                    
                    if not found:
                         combined = " ".join([(e.text or "") for e in uf_spans]).upper()
                         for sig in ESTADOS:
                             if re.search(r'\b' + re.escape(sig) + r'\b', combined):
                                 found = sig
                                 break
                    
                    linha_atual[1].value = found if found else 'ND'
                except:
                    linha_atual[1].value = 'Erro UF'

                # Fechar detalhe
                try:
                    time.sleep(0.2)
                    fechar = None
                    # Tenta seletores de fechar
                    try:
                        fechar = driver.find_element(By.CSS_SELECTOR, "button.button.s-cancel")
                    except:
                        try:
                            fechar = driver.find_element(By.XPATH, "//button[contains(concat(' ', normalize-space(@class), ' '), ' s-cancel ') and contains(., 'Cancelar')]")
                        except:
                            fechar = None
                    
                    if fechar:
                        click_element_retry(fechar, attempts=3, pause=0.2)
                        time.sleep(0.25)
                except:
                    pass

                # Marca processado
                try:
                    driver.execute_script("arguments[0].setAttribute('data-processed','1')", el)
                except:
                    pass

                idx += 1
                processed += 1
            
            # Salva parcial
            wb.save(EXCEL_PATH)
            percent = 50 + int(45 * (count_detalhado / total_detalhar))
            log(f"📝 Evento {evento} concluído ({count_detalhado}/{total_detalhar})", percent) # type: ignore

        # ===============================
        # ORDENAÇÃO FINAL
        # ===============================
        log("Sort e Salvando final...", 95)
        try:
            # Recarrega para ordenar
            wb = load_workbook(EXCEL_PATH)
            ws = wb["Eventos"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            def sort_key(row):
                v = row[0]
                if v is None: return (1, "")
                s = str(v).strip()
                try:
                    return (0, int(s))
                except:
                    return (1, s.lower())

            rows_sorted = sorted(rows, key=sort_key)

            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
                
            for r in rows_sorted:
                ws.append(list(r))
            
            wb.save(EXCEL_PATH)
        except Exception as e:
            log(f"⚠️ Erro na ordenação final: {e}")

        driver.quit()
        log(f"✅ Processo finalizado com sucesso! Arquivo: {filename}", 100)
        
        return {
            "sucesso": True,
            "arquivo": filename,
            "eventos_coletados": total_detalhar,
            "logs": f"Sucesso. {total_detalhar} eventos processados."
        }

    except Exception as e:
        if 'driver' in locals():
            driver.quit()
        error_msg = f"Erro fatal na execução: {e}"
        log(f"❌ {error_msg}")
        raise e

# ===============================
# FUNÇÃO ASYNC WRAPPER
# ===============================
EXECUCOES_ATIVAS = {}

def executar_robo_vale_async(data_coleta, user_id, log_callback=None):
    """Executa o robô do Vale em thread separada e isolada por usuário."""
    app = create_app()

    if user_id in EXECUCOES_ATIVAS:
        print(f"⚠️ Robô já em execução para user_id={user_id}. Ignorando nova chamada.")
        return

    def tarefa():
        temp_dir = tempfile.mkdtemp(prefix=f"vale_{str(user_id)[:6]}_")
        os.environ["WDM_LOCAL"] = temp_dir 
        print(f"🧵 [Thread-{str(user_id)[:6]}] Iniciando robô no diretório {temp_dir}")
        
        success = False
        try:
            with app.app_context():
                EXECUCOES_ATIVAS[user_id] = True
                executar_robo_vale(
                    data_coleta=data_coleta,
                    user_id=user_id,
                    log_callback=log_callback
                )
                success = True
        except Exception as e:
            print(f"❌ [Thread-{str(user_id)[:6]}] Erro: {e}")
        finally:
            EXECUCOES_ATIVAS.pop(user_id, None)
            shutil.rmtree(temp_dir, ignore_errors=True)
            print(f"🧹 [Thread-{str(user_id)[:6]}] Finalizou (Sucesso: {success})")

    thread = threading.Thread(target=tarefa, daemon=True)
    thread.start()


